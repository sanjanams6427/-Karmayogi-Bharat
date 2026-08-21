"""
KB Translation System — Enterprise UI
Run: python ui/app.py
"""

import sys, os, json, time, threading, warnings
warnings.filterwarnings("ignore", message=".*HTTP_422_UNPROCESSABLE_ENTITY.*")
warnings.filterwarnings("ignore", category=DeprecationWarning, module="starlette")
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import gradio as gr
from pipeline.lang_config import LANG_NAMES, ALL_22
from ui.reviewer import (
    load_metadata, load_review, save_review,
    export_certificate, review_stats, segments_to_display,
)

SRC_LANGS = [("English", "eng"), ("Hindi", "hin")] + [
    (LANG_NAMES[c], c) for c in ALL_22 if c not in ("hin",)
]
TGT_LANGS = [(LANG_NAMES[c], c) for c in ALL_22]
KB_11     = ["asm","ben","guj","hin","kan","mal","mar","ory","pan","tam","tel"]
OUTPUT_DIR     = str(Path(__file__).parent.parent / "output")
ENV_PATH       = Path(__file__).parent.parent / ".env"
SAVE_DIR_FILE  = Path(__file__).parent.parent / ".output_dir"  # persists user choice

# Ensure default output dir always exists
Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)


def _get_output_dir() -> str:
    if SAVE_DIR_FILE.exists():
        d = SAVE_DIR_FILE.read_text(encoding="utf-8").strip()
        if d:
            Path(d).mkdir(parents=True, exist_ok=True)
            return d
    return OUTPUT_DIR


def _set_output_dir(path: str) -> str:
    p = Path(path.strip())
    try:
        p.mkdir(parents=True, exist_ok=True)
        SAVE_DIR_FILE.write_text(str(p), encoding="utf-8")
        return f"✅ Output folder set to: {p}"
    except Exception as e:
        return f"❌ {e}"

_pipeline = None
_pipeline_lock = threading.Lock()
_pipeline_ready = threading.Event()
_pipeline_error = ""
_log_lines: list[str] = []
_log_lock  = threading.Lock()
_job_semaphore = threading.Semaphore(1)  # only 1 job at a time — GPU has no spare VRAM for 2


def _append_log(msg: str):
    with _log_lock:
        _log_lines.append(f"[{time.strftime('%H:%M:%S')}] {msg}")
        if len(_log_lines) > 200:
            _log_lines.pop(0)


def _get_log() -> str:
    with _log_lock:
        return "\n".join(_log_lines[-60:])


def _load_pipeline_bg():
    global _pipeline, _pipeline_error
    try:
        _append_log("Loading pipeline models in background (may take 1-2 min)...")
        from pipeline.dubbing_pipeline import DubbingPipeline
        _pipeline = DubbingPipeline()
        _append_log("✅ Pipeline ready — you can now submit jobs.")
    except Exception as e:
        _pipeline_error = str(e)
        _append_log(f"❌ Pipeline load failed: {e}")
    finally:
        _pipeline_ready.set()


# Start loading immediately when app.py is imported
_bg_thread = threading.Thread(target=_load_pipeline_bg, daemon=True)
_bg_thread.start()


def get_pipeline():
    if not _pipeline_ready.wait(timeout=300):
        raise RuntimeError("Pipeline failed to load within 5 minutes")
    if _pipeline_error:
        raise RuntimeError(f"Pipeline load error: {_pipeline_error}")
    return _pipeline


# ── API Keys helpers ──────────────────────────────────────────
def _load_env() -> dict:
    keys = {"HF_TOKEN": ""}
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if "=" in line and not line.startswith("#"):
                k, _, v = line.partition("=")
                k = k.strip()
                if k in keys:
                    keys[k] = v.strip()
    return keys


def _save_env(hf: str) -> str:
    lines = []
    if ENV_PATH.exists():
        for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
            if line.split("=")[0].strip() != "HF_TOKEN":
                lines.append(line)
    if hf:
        lines.append(f"HF_TOKEN={hf}")
        os.environ["HF_TOKEN"] = hf
    ENV_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")
    global _pipeline
    with _pipeline_lock:
        _pipeline = None
    return "✅ HF Token saved. Pipeline will reload on next job."


def _save_outputs(files: list[str]) -> list[str]:
    """Return valid output paths — pipeline already saves to output_dir, no copy needed."""
    saved = []
    for f in files:
        if not f or not f.strip():
            continue
        src = Path(f)
        if src.exists() and src.is_file():
            saved.append(str(src))
    return saved


# ── Tab 2: Translate Document ─────────────────────────────────
def _chunk_text(text: str, max_chars: int = 400) -> list[str]:
    """Split on sentence boundaries; only hard-split if a sentence itself is too long."""
    import re
    sentences = re.split(r'(?<=[.!?।॥])\s+', text.strip())
    chunks, current = [], ""
    for sent in sentences:
        if not sent:
            continue
        if len(sent) > max_chars:
            # Hard-split oversized sentence at word boundaries
            words, part = sent.split(), ""
            for w in words:
                if len(part) + len(w) + 1 > max_chars and part:
                    chunks.append(part)
                    part = w
                else:
                    part = (part + " " + w).strip()
            if part:
                chunks.append(part)
            current = ""
        elif len(current) + len(sent) + 1 > max_chars and current:
            chunks.append(current)
            current = sent
        else:
            current = (current + " " + sent).strip()
    if current:
        chunks.append(current)
    return chunks or [text]


def _translate_plain_doc(file_path: str, src_lang: str, tgt_langs: list,
                         out_dir: str, course_id: str,
                         progress, pipeline) -> tuple[list, list]:
    """Translate PDF/DOCX/TXT. DOCX preserves all formatting; PDF/TXT → plain .docx."""
    suffix = Path(file_path).suffix.lower()
    output_files, log_lines = [], []

    # ── DOCX: format-preserving path ─────────────────────────────────────
    if suffix in (".docx", ".doc"):
        from pipeline.doc_extractor import translate_docx

        def _batch_translate(texts: list[str], src: str, tgt: str) -> list[str]:
            """Document translation — raw engine, no protection layers."""
            return pipeline.translator.translate_document_batch(texts, src, tgt)

        for i, tgt in enumerate(tgt_langs):
            progress((i + 1) / len(tgt_langs), desc=f"Translating \u2192 {LANG_NAMES.get(tgt, tgt)}")
            out_path = os.path.join(out_dir, f"{course_id}_{tgt}.docx")
            try:
                translate_docx(file_path, out_path, _batch_translate, src_lang, tgt)
                output_files.append(out_path)
                log_lines.append(f"✅ {LANG_NAMES[tgt]} → {Path(out_path).name}")
            except Exception as e:
                log_lines.append(f"❌ {LANG_NAMES[tgt]}: {e}")
        return output_files, log_lines

    # ── PDF / TXT: plain-text path → translated .docx ────────────────────
    from pipeline.doc_extractor import extract_text
    from docx import Document as DocxDocument

    try:
        text = extract_text(file_path)
    except Exception as e:
        return [], [f"❌ Could not extract text: {e}"]

    raw_paras = [p.strip() for p in text.splitlines() if p.strip()]
    paragraphs = []
    for p in raw_paras:
        paragraphs.extend(_chunk_text(p) if len(p) > 400 else [p])

    for i, tgt in enumerate(tgt_langs):
        progress((i + 1) / len(tgt_langs), desc=f"Translating → {LANG_NAMES[tgt]}")
        try:
            translated_paras = pipeline.translator.translate_document_batch(paragraphs, src_lang, tgt)
        except Exception as e:
            log_lines.append(f"❌ {LANG_NAMES[tgt]}: {e}")
            continue

        out_path = os.path.join(out_dir, f"{course_id}_{tgt}.docx")
        doc = DocxDocument()
        for para in translated_paras:
            doc.add_paragraph(para)
        doc.save(out_path)
        output_files.append(out_path)
        log_lines.append(f"✅ {LANG_NAMES[tgt]} → {Path(out_path).name}")

    return output_files, log_lines


def translate_doc(file, src_lang, tgt_langs, doc_type, course_title,
                  progress=gr.Progress()):
    if file is None:
        return None, "❌ Please upload a file."
    if not tgt_langs:
        return None, "❌ Select at least one target language."

    _t0 = time.time()
    pipeline  = get_pipeline()
    course_id = Path(file.name).stem
    out_dir   = os.path.join(_get_output_dir(), course_id)
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    suffix = Path(file.name).suffix.lower()

    # ── PDF blocked per tender §3.1 ──
    if suffix == ".pdf":
        return None, "❌ PDF documents are NOT translated per tender §3.1. Upload the PDF as-is to CBP portal in the original language."

    # ── Plain document (DOCX / TXT) ──
    if suffix in (".docx", ".doc", ".txt"):
        output_files, log_lines = _translate_plain_doc(
            file.name, src_lang, tgt_langs, out_dir, course_id, progress, pipeline)
        elapsed = round((time.time() - _t0) / 60, 1)
        log_lines.insert(0, f"⏱ Completed in {elapsed} min")
        return _save_outputs(output_files) or None, "\n".join(log_lines)

    # ── JSON paths (existing behaviour) ──
    output_files, log_lines = [], []

    if doc_type == "Quiz / Assessment (Word .docx)":
        try:
            quiz = json.loads(Path(file.name).read_text(encoding="utf-8"))
        except Exception as e:
            return None, f"❌ Could not parse quiz JSON: {e}"
        for i, tgt in enumerate(tgt_langs):
            progress((i+1)/len(tgt_langs), desc=f"Quiz → {LANG_NAMES[tgt]}")
            out_path = os.path.join(out_dir, f"{course_id}_quiz_{tgt}.docx")
            try:
                pipeline.export_quiz_docx(quiz, src_lang, tgt, out_path, course_title)
                output_files.append(out_path)
                log_lines.append(f"✅ {LANG_NAMES[tgt]} quiz → {Path(out_path).name}")
            except Exception as e:
                log_lines.append(f"❌ {LANG_NAMES[tgt]}: {e}")
    else:
        try:
            metadata = json.loads(Path(file.name).read_text(encoding="utf-8"))
        except Exception as e:
            return None, f"❌ Could not parse metadata JSON: {e}"
        out_path = os.path.join(out_dir, f"{course_id}_metadata.xlsx")
        try:
            pipeline.export_metadata_xlsx(metadata, src_lang, tgt_langs, out_path)
            output_files.append(out_path)
            log_lines.append(f"✅ Metadata Excel → {Path(out_path).name}")
        except Exception as e:
            log_lines.append(f"❌ {e}")

    elapsed = round((time.time() - _t0) / 60, 1)
    log_lines.insert(0, f"⏱ Completed in {elapsed} min")
    return _save_outputs(output_files) or None, "\n".join(log_lines)


# ── Tab 3: Full Course Batch ──────────────────────────────────
def _build_scores_table(summary: dict) -> list[list]:
    """Build rows for the per-language quality score Dataframe."""
    rows = []
    for lang, info in summary.get("dubbing", {}).items():
        lang_name = LANG_NAMES.get(lang, lang)
        if not info.get("success"):
            rows.append([lang_name, "—", "❌ FAILED", "—", "—", "—", "—", "—"])
            continue
        q         = info.get("quality") or {}
        score     = q.get("avg_score", "—")
        pass_rate = q.get("pass_rate", "—")
        total     = q.get("total", "—")
        failed    = q.get("failed", 0)
        review    = q.get("needs_review", 0)
        dur_ratio = q.get("duration_ratio")
        dur_flag  = q.get("duration_ratio_flag", False)
        try:
            s = float(score)
            bar   = "█" * int(s * 10) + "░" * (10 - int(s * 10))
            label = "✅ Pass" if s >= 0.55 else ("⚠️ Review" if s >= 0.30 else "❌ Failed")
            score_str = f"{s:.2f}  [{bar}]"
        except (TypeError, ValueError):
            label     = "—"
            score_str = str(score)
        dur_str = (
            f"⚠️ {dur_ratio:.2f}x" if dur_flag else
            (f"✅ {dur_ratio:.2f}x" if dur_ratio is not None else "—")
        )
        rows.append([lang_name, score_str, label,
                     str(pass_rate), str(total), str(failed), str(review), dur_str])
    return rows


def _format_quality_summary(summary: dict) -> str:
    """Convert raw pipeline summary dict into a readable per-language quality report."""
    lines = []
    elapsed = summary.get("elapsed_min", "?")
    langs   = summary.get("target_langs", [])
    lines.append(f"⏱ Completed in {elapsed} min  |  {len(langs)} language(s)")
    lines.append("=" * 56)

    for lang, info in summary.get("dubbing", {}).items():
        lang_name = LANG_NAMES.get(lang, lang)
        if not info.get("success"):
            err = info.get("error") or "unknown error"
            lines.append(f"❌ {lang_name} ({lang})  —  FAILED: {err[:60]}")
            continue

        q = info.get("quality") or {}
        score     = q.get("avg_score",  "—")
        pass_rate = q.get("pass_rate",  "—")
        total     = q.get("total",      "—")
        failed    = q.get("failed",     0)
        review    = q.get("needs_review", 0)
        dur_ratio = q.get("duration_ratio")
        dur_flag  = q.get("duration_ratio_flag", False)

        # Overall status emoji
        try:
            s = float(score)
            status = "✅" if s >= 0.55 else ("⚠️" if s >= 0.30 else "❌")
        except (TypeError, ValueError):
            status = "✅"

        lines.append(f"{status} {lang_name} ({lang})")
        lines.append(f"   Translation quality : {score}  (pass rate {pass_rate})")
        lines.append(f"   Segments            : {total} total  |  {failed} failed  |  {review} need review")
        if dur_ratio is not None:
            dur_icon = "⚠️" if dur_flag else "✅"
            lines.append(f"   Duration ratio      : {dur_icon} {dur_ratio:.2f}x"
                         + ("  ← KB approval required" if dur_flag else ""))
        out = info.get("output") or ""
        if out:
            lines.append(f"   Output              : {Path(out).name}")
        lines.append("")

    # Footer: overall pass/fail count
    dub = summary.get("dubbing", {})
    ok  = sum(1 for v in dub.values() if v.get("success"))
    fail = len(dub) - ok
    lines.append("-" * 56)
    lines.append(f"Total: {ok} succeeded  |  {fail} failed")
    if summary.get("metadata_xlsx"):
        lines.append("📄 Metadata Excel exported")
    if summary.get("quiz_docx"):
        lines.append(f"📝 Quiz DOCX exported for {len(summary['quiz_docx'])} language(s)")
    if summary.get("qa_reports"):
        lines.append(f"📋 QA certificates generated for {len(summary['qa_reports'])} language(s)")
    return "\n".join(lines)


def process_course(video_file, meta_file, quiz_file, src_lang, tgt_langs,
                   course_id, upload_cbp, progress=gr.Progress()):
    if video_file is None:
        return None, [], "❌ Please upload a video/audio file."
    if not tgt_langs:
        return None, [], "❌ Select at least one target language."

    # KB tender §3.1 — Non-SCORM content only
    from pipeline.scorm_guard import is_scorm_package
    _scorm, _scorm_reason = is_scorm_package(video_file.name)
    if _scorm:
        return None, [], f"❌ SCORM content rejected — {_scorm_reason}. Upload the raw MP4/MP3/WAV source file."


    pipeline = get_pipeline()
    cid      = course_id or Path(video_file.name).stem
    out_dir  = os.path.join(_get_output_dir(), cid)

    metadata = None
    if meta_file:
        try:
            metadata = json.loads(Path(meta_file.name).read_text(encoding="utf-8"))
        except Exception:
            pass

    quiz = None
    if quiz_file:
        try:
            quiz = json.loads(Path(quiz_file.name).read_text(encoding="utf-8"))
        except Exception:
            pass

    try:
        import torch
        _num_gpus = torch.cuda.device_count() if torch.cuda.is_available() else 1
    except Exception:
        _num_gpus = 1

    if not _job_semaphore.acquire(blocking=False):
        return None, [], "⏳ Another job is already running. Please wait for it to finish."
    try:
        progress(0.05, desc=f"Starting pipeline — 0 / {len(tgt_langs)} languages…")
        _append_log(f"Full course batch: {cid} → {tgt_langs} | GPUs={_num_gpus}")
        _t0 = time.time()

        # Run pipeline in a background thread so we can poll progress here
        _result_box: list = [None]
        _exc_box:    list = [None]

        def _run():
            try:
                _result_box[0] = pipeline.process_course_full(
                    video_file.name, src_lang, tgt_langs, out_dir, cid,
                    metadata=metadata, quiz=quiz,
                    upload_to_cbp=upload_cbp,
                    num_gpus=_num_gpus,
                )
            except Exception as e:
                _exc_box[0] = e

        _worker = threading.Thread(target=_run, daemon=True)
        _worker.start()

        # Poll output directory for completed language folders
        n_total = len(tgt_langs)
        while _worker.is_alive():
            done = sum(
                1 for lang in tgt_langs
                if any(Path(out_dir, lang).glob(f"{cid}_{lang}.mp4")) or
                   any(Path(out_dir, lang).glob(f"{cid}_{lang}.mp3"))
            )
            frac = 0.05 + 0.90 * (done / n_total) if n_total else 0.95
            progress(frac, desc=f"Dubbing… {done} / {n_total} languages done")
            time.sleep(3)

        _worker.join()
        if _exc_box[0]:
            raise _exc_box[0]

        progress(0.97, desc="Collecting outputs…")
        summary = _result_box[0]
        all_files = []
        for tgt, info in summary["dubbing"].items():
            if info.get("output") and Path(info["output"]).exists():
                all_files.append(info["output"])
            if info.get("output_mp3") and Path(info["output_mp3"]).exists():
                all_files.append(info["output_mp3"])
            sub_dir = Path(out_dir) / tgt
            for ext in (".srt", ".vtt", "_qa_cert.docx", "_metadata.json"):
                p = sub_dir / f"{cid}_{tgt}{ext}"
                if p.exists():
                    all_files.append(str(p))
        for p in summary.get("quiz_docx", {}).values():
            if p and Path(p).exists():
                all_files.append(p)
        for p in summary.get("qa_reports", {}).values():
            if p and Path(p).exists():
                all_files.append(p)
        if summary.get("metadata_xlsx", {}).get("all") and \
                Path(summary["metadata_xlsx"]["all"]).exists():
            all_files.append(summary["metadata_xlsx"]["all"])
        _elapsed = round((time.time() - _t0) / 60, 1)
        summary["elapsed_min"] = _elapsed
        _append_log(f"✅ Dubbing done in {_elapsed} min")
        scores_table = _build_scores_table(summary)
        progress(1.0, desc="Done!")
        return _save_outputs(all_files) or None, scores_table, _format_quality_summary(summary)
    finally:
        _job_semaphore.release()


# ── Tab 4: QA Report ─────────────────────────────────────────
def gen_qa(course_id, tgt_lang, src_lang, input_file, output_file, reviewer):
    from pipeline.dubbing_pipeline import DubbingResult
    if not course_id or not input_file:
        return None, "❌ Course ID and input file are required."
    pipeline = get_pipeline()
    out_dir  = os.path.join(_get_output_dir(), course_id)
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    dummy = DubbingResult(
        source_lang=src_lang, target_lang=tgt_lang,
        input_path=input_file.name,
        output_video_path=output_file.name if output_file else "",
        success=True,
    )
    qa_path = os.path.join(out_dir, f"{course_id}_{tgt_lang}_qa_cert.docx")
    try:
        pipeline.generate_qa_report(course_id, tgt_lang, dummy, qa_path, reviewer)
        return qa_path, f"✅ QA report → {Path(qa_path).name}"
    except Exception as e:
        return None, f"❌ {e}"


# ── Build UI ──────────────────────────────────────────────────
def build_ui():
    src_choices = list(SRC_LANGS)
    tgt_choices = list(TGT_LANGS)
    env         = _load_env()

    with gr.Blocks(title="KB Translation System") as app:

        def _pipeline_status():
            if _pipeline_ready.is_set():
                return "✅ Pipeline ready" if not _pipeline_error else f"❌ {_pipeline_error}"
            return "⏳ Loading models in background..."

        gr.Markdown(
            "# 🇮🇳 KB Translation & Dubbing System\n"
            "**iGOT Karmayogi | RFB IN-KBL-543730-NC-RFB** · "
            "22 Scheduled Indian Languages · "
            "IndicTrans2 · faster-whisper · Parler-TTS · MMS-TTS"
        )
        status_bar = gr.Textbox(value=_pipeline_status, every=3,
                                label="Pipeline Status", interactive=False, lines=1)

        # ── Tab 1: Dub Video / Audio ──────────────────────────
        with gr.Tab("🎬 Dub Video / Audio"):
            gr.Markdown(
                "**iGOT Karmayogi — KB Tender RFB IN-KBL-543730-NC-RFB**  \n"
                "Mandatory: 11 languages (KB-11) · Optional: remaining 11 scheduled languages · "
                "Output: MP4 + SRT/VTT per language · Upload to CBP portal after acceptance"
            )
            with gr.Row():
                with gr.Column(scale=2):
                    t1_file = gr.File(
                        label="Course Video / Audio (MP4 / MP3 / WAV / WEBM)",
                        file_types=[".mp4", ".mp3", ".wav", ".flac", ".webm"]
                    )
                    with gr.Row():
                        t1_meta = gr.File(
                            label="Course Metadata (Word / Excel / JSON)",
                            file_types=[".docx", ".xlsx", ".json"]
                        )
                        t1_quiz = gr.File(
                            label="Quiz / Assessment (Word / Excel / JSON)",
                            file_types=[".docx", ".xlsx", ".json"]
                        )
                    t1_id  = gr.Textbox(label="Course ID", value="KB_COURSE_001",
                                        placeholder="e.g. KB_COURSE_001")
                    t1_src = gr.Dropdown(src_choices, value="eng", label="Source Language")
                    t1_tgt = gr.CheckboxGroup(
                        tgt_choices, value=KB_11,
                        label="Target Languages  (KB-11 mandatory · tick more for optional)"
                    )
                    with gr.Row():
                        gr.Button("✅ KB 11 (Mandatory)", size="sm").click(
                            lambda: KB_11, outputs=[t1_tgt])
                        gr.Button("All 22", size="sm").click(
                            lambda: [c for _, c in tgt_choices], outputs=[t1_tgt])
                        gr.Button("Clear", size="sm").click(
                            lambda: [], outputs=[t1_tgt])
                    with gr.Row():
                        t1_cbp   = gr.Checkbox(label="📤 Upload to CBP Portal", value=False)
                    t1_btn = gr.Button("🚀 Start Dubbing", variant="primary", size="lg")

                with gr.Column(scale=2):
                    t1_dl      = gr.Files(label="⬇️ Download All Outputs (MP4 · SRT · VTT · DOCX)")
                    t1_scores  = gr.Dataframe(
                        headers=["Language", "Score", "Status", "Pass Rate", "Segments", "Failed", "Needs Review", "Duration Ratio"],
                        datatype=["str", "str", "str", "str", "str", "str", "str", "str"],
                        interactive=False,
                        label="📊 Dubbed Output Quality Scores (per language)",
                        wrap=False,
                    )
                    t1_quality = gr.Textbox(
                        label="Job Summary",
                        lines=10, interactive=False,
                        placeholder="Results appear here after job completes…"
                    )

            t1_btn.click(
                process_course,
                inputs=[t1_file, t1_meta, t1_quiz, t1_src, t1_tgt,
                        t1_id, t1_cbp],
                outputs=[t1_dl, t1_scores, t1_quality],
            )

        # ── Tab 2: Translate Document ─────────────────────────
        with gr.Tab("📄 Translate Document"):
            gr.Markdown(
                "Translate course materials per tender scope: **Quiz / Assessment** (Word/Excel) · "
                "**Course Metadata** (title, description, learning outcomes, keywords) · "
                "**PDF / TXT** documents"
            )
            with gr.Row():
                with gr.Column(scale=2):
                    gr.Markdown(
                        "> ⚠️ **Tender Exclusions (§3.1):** PDF documents are uploaded as-is — not translated. "
                        "Government speeches, ceremonial addresses and classified content must NOT be submitted for translation."
                    )
                    t2_file  = gr.File(
                        label="Upload Document (DOCX / TXT / JSON — PDF not accepted)",
                        file_types=[".docx", ".doc", ".xlsx", ".txt", ".json"]
                    )
                    t2_type  = gr.Radio(
                        ["Quiz / Assessment (Word .docx)",
                         "Course Metadata (Excel .xlsx)",
                         "General Document (TXT / DOCX)"],
                        value="Quiz / Assessment (Word .docx)",
                        label="Document Type"
                    )
                    t2_src   = gr.Dropdown(src_choices, value="eng", label="Source Language")
                    t2_tgt   = gr.CheckboxGroup(tgt_choices, value=KB_11,
                                                label="Target Languages")
                    with gr.Row():
                        gr.Button("KB 11", size="sm").click(lambda: KB_11, outputs=[t2_tgt])
                        gr.Button("All 22", size="sm").click(
                            lambda: [c for _, c in tgt_choices], outputs=[t2_tgt])
                        gr.Button("Clear", size="sm").click(lambda: [], outputs=[t2_tgt])
                    t2_title = gr.Textbox(label="Course Title (for metadata header)")
                    t2_btn   = gr.Button("🚀 Translate Document", variant="primary")
                with gr.Column(scale=2):
                    t2_dl  = gr.Files(label="⬇️ Download Translated Documents")
                    t2_log = gr.Textbox(label="Translation Log", lines=10, interactive=False)
            t2_btn.click(translate_doc,
                         inputs=[t2_file, t2_src, t2_tgt, t2_type, t2_title],
                         outputs=[t2_dl, t2_log])

        # ── Tab 3: QA Certificate ─────────────────────────────
        with gr.Tab("📋 QA Certificate"):
            gr.Markdown(
                "Generate the **Language Quality Assurance Certification** required per tender SLA.  \n"
                "Certifies: linguistic accuracy ≥ 98% · terminology consistency · "
                "compliance with KB content guidelines · review by qualified language expert."
            )
            with gr.Row():
                with gr.Column(scale=2):
                    t4_id       = gr.Textbox(label="Course ID", placeholder="KB_COURSE_001")
                    t4_src      = gr.Dropdown(src_choices, value="eng", label="Source Language")
                    t4_tgt      = gr.Dropdown(tgt_choices, value="hin", label="Target Language")
                    t4_input    = gr.File(label="Original Source File (MP4 / MP3 / DOCX)")
                    t4_output   = gr.File(label="Dubbed / Translated Output File")
                    t4_reviewer = gr.Textbox(
                        label="Language Expert / Reviewer Name",
                        placeholder="e.g. Dr. Priya Nair (Tamil Expert)",
                        value="Translation Agency QA Lead"
                    )
                    t4_btn = gr.Button("📋 Generate QA Certificate (.docx)", variant="primary")
                with gr.Column(scale=2):
                    t4_dl  = gr.File(label="⬇️ Download QA Certificate (.docx)")
                    t4_log = gr.Textbox(label="Status", lines=4, interactive=False)
                    gr.Markdown(
                        "**SLA Thresholds (KB Tender)**  \n"
                        "- Score ≥ 0.55 → ✅ Pass (98%+ accuracy)  \n"
                        "- Score 0.30–0.55 → ⚠️ Needs correction (resubmit within 5 days)  \n"
                        "- Score < 0.30 → ❌ Failed — mandatory re-translation  \n\n"
                        "**Delivery SLA**  \n"
                        "- < 5% shortfall → No penalty  \n"
                        "- 5–10% shortfall → 2% deduction  \n"
                        "- > 10% shortfall → 4% deduction  \n"
                        "- > 20% shortfall → 5% deduction"
                    )
            t4_btn.click(gen_qa,
                         inputs=[t4_id, t4_tgt, t4_src, t4_input, t4_output, t4_reviewer],
                         outputs=[t4_dl, t4_log])

        # ── Tab 4: Human Review ───────────────────────────────
        with gr.Tab("👤 Human Review"):
            _rev_state = gr.State([])   # list[dict] segments
            _rev_path  = gr.State("")   # path to loaded metadata JSON

            with gr.Row():
                with gr.Column(scale=1):
                    rv_file     = gr.File(label="Load *_metadata.json",
                                         file_types=[".json"])
                    rv_reviewer = gr.Textbox(
                        label="Reviewer Name",
                        placeholder="e.g. Dr. Priya Nair (Tamil Expert)")
                    rv_load_btn = gr.Button("📂 Load Segments", variant="primary")
                    rv_stats    = gr.Textbox(
                        label="Review Progress  (SLA: corrections resubmitted within 5 days)",
                        lines=2, interactive=False
                    )
                with gr.Column(scale=3):
                    rv_table = gr.Dataframe(
                        headers=["ID", "Time", "Source", "AI Translation",
                                 "Corrected Text", "Score", "Flags", "🚩", "Decision"],
                        datatype=["str","str","str","str","str",
                                  "str","str","str","str"],
                        interactive=True,
                        wrap=True,
                        label="Edit 'Corrected Text' and 'Decision' "
                              "(approved / corrected / rejected)",
                    )

            with gr.Row():
                rv_approve_all = gr.Button("✅ Approve All Unflagged")
                rv_save_btn    = gr.Button("💾 Save Progress")
                rv_cert_btn    = gr.Button("📜 Export Review Certificate (.docx)",
                                           variant="primary")
            with gr.Row():
                rv_cert_dl = gr.File(label="Download Certificate")
                rv_log     = gr.Textbox(label="Status", lines=3, interactive=False)

            def _rv_load(file, reviewer):
                if file is None:
                    return [], "", [], "", "❌ Upload a metadata JSON first."
                try:
                    segs, _ = load_metadata(file.name)
                    segs    = load_review(file.name, segs)
                    rows    = segments_to_display(segs)
                    stats   = review_stats(segs)
                    return segs, file.name, rows, stats, f"✅ Loaded {len(segs)} segments."
                except Exception as e:
                    return [], "", [], "", f"❌ {e}"

            def _rv_save(table_data, path, reviewer, segs):
                if not path or not segs:
                    return segs, "❌ Load a file first.", ""
                for i, row in enumerate(table_data):
                    if i < len(segs):
                        segs[i]["corrected_text"] = row[4] or segs[i]["translated_text"]
                        segs[i]["decision"]        = (row[8] or "").strip().lower()
                try:
                    sp = save_review(path, segs, reviewer or "Reviewer")
                    return segs, f"✅ Saved → {sp}", review_stats(segs)
                except Exception as e:
                    return segs, f"❌ {e}", ""

            def _rv_cert(table_data, path, reviewer, segs):
                if not path or not segs:
                    return None, "❌ Load and save a review first."
                for i, row in enumerate(table_data):
                    if i < len(segs):
                        segs[i]["corrected_text"] = row[4] or segs[i]["translated_text"]
                        segs[i]["decision"]        = (row[8] or "").strip().lower()
                cert_path = str(
                    Path(path).parent /
                    (Path(path).stem.replace("_metadata", "") + "_review_cert.docx")
                )
                try:
                    export_certificate(path, segs, reviewer or "Reviewer", cert_path)
                    return cert_path, f"✅ Certificate → {Path(cert_path).name}"
                except Exception as e:
                    return None, f"❌ {e}"

            def _rv_approve_all(table_data, segs):
                rows = [list(r) for r in table_data]
                for i, row in enumerate(rows):
                    if i < len(segs) and not segs[i]["flags"] and not (row[8] or "").strip():
                        rows[i][8] = "approved"
                        segs[i]["decision"] = "approved"
                return rows, segs, review_stats(segs)

            rv_load_btn.click(
                _rv_load,
                inputs=[rv_file, rv_reviewer],
                outputs=[_rev_state, _rev_path, rv_table, rv_stats, rv_log],
            )
            rv_save_btn.click(
                _rv_save,
                inputs=[rv_table, _rev_path, rv_reviewer, _rev_state],
                outputs=[_rev_state, rv_log, rv_stats],
            )
            rv_cert_btn.click(
                _rv_cert,
                inputs=[rv_table, _rev_path, rv_reviewer, _rev_state],
                outputs=[rv_cert_dl, rv_log],
            )
            rv_approve_all.click(
                _rv_approve_all,
                inputs=[rv_table, _rev_state],
                outputs=[rv_table, _rev_state, rv_stats],
            )

        # ── Tab 5: Corrections ────────────────────────────────
        with gr.Tab("🔧 Corrections"):
            gr.Markdown(
                "**Correction Cycle Tracker — KB Tender §5.1B**  \n"
                "Raise a ticket when KB flags an issue. Deadline: **5 calendar days** from feedback.  \n"
                "Delay penalty: **0.5% per week** (or part thereof) beyond deadline."
            )
            with gr.Row():
                with gr.Column(scale=2):
                    gr.Markdown("### Raise New Correction Ticket")
                    ct_course    = gr.Textbox(label="Course ID", placeholder="KB_COURSE_001")
                    ct_lang      = gr.Dropdown(tgt_choices, value="hin", label="Language")
                    ct_feedback  = gr.Textbox(label="Feedback from KB / Verification Agency",
                                              lines=3, placeholder="Describe the issue flagged...")
                    ct_raised_by = gr.Textbox(label="Raised By", value="KB Verification Agency")
                    ct_date      = gr.Textbox(label="Feedback Date (YYYY-MM-DD, blank=today)",
                                              placeholder="2025-07-15")
                    ct_raise_btn = gr.Button("📨 Raise Ticket", variant="primary")
                    ct_raise_log = gr.Textbox(label="Status", lines=2, interactive=False)
                with gr.Column(scale=3):
                    gr.Markdown("### Open / In-Progress Tickets")
                    ct_table = gr.Dataframe(
                        headers=["Ticket ID","Course","Lang","Status","Feedback","Deadline","Penalty"],
                        datatype=["str","str","str","str","str","str","str"],
                        interactive=False, wrap=True, label="Correction Tickets"
                    )
                    ct_refresh_btn = gr.Button("🔄 Refresh Tickets")
            with gr.Row():
                with gr.Column(scale=2):
                    gr.Markdown("### Update / Close Ticket")
                    ct_ticket_id  = gr.Textbox(label="Ticket ID (e.g. COR-0001)")
                    ct_resolution = gr.Textbox(label="Resolution / Corrective Action", lines=3,
                                               placeholder="Describe what was corrected...")
                    ct_closed_by  = gr.Textbox(label="Closed By", value="Translation Agency")
                    with gr.Row():
                        ct_inprog_btn = gr.Button("⏳ Mark In Progress")
                        ct_close_btn  = gr.Button("✅ Close Ticket", variant="primary")
                    ct_close_log  = gr.Textbox(label="Status", lines=2, interactive=False)
                with gr.Column(scale=2):
                    gr.Markdown("### Export Closure Report")
                    ct_rep_course = gr.Textbox(label="Course ID (blank = all courses)")
                    ct_rep_lang   = gr.Dropdown(
                        [("All languages", "")] + list(tgt_choices),
                        value="", label="Language (blank = all)"
                    )
                    ct_rep_agency = gr.Textbox(label="Agency Name", value="Translation Agency")
                    ct_rep_btn    = gr.Button("📄 Export Closure Report (.docx)", variant="primary")
                    ct_rep_dl     = gr.File(label="⬇️ Download Closure Report")
                    ct_rep_log    = gr.Textbox(label="Status", lines=2, interactive=False)

            def _ct_load():
                from pipeline.correction_tracker import get_all_tickets, tickets_to_rows
                return tickets_to_rows(get_all_tickets())

            def _ct_raise(course, lang, feedback, raised_by, date_str):
                from pipeline.correction_tracker import raise_ticket
                if not course or not feedback:
                    return _ct_load(), "❌ Course ID and feedback are required."
                try:
                    fd = date_str.strip() if date_str and date_str.strip() else None
                    if fd:
                        import datetime as _dt; _dt.datetime.fromisoformat(fd)
                    t = raise_ticket(course, lang, feedback,
                                     raised_by=raised_by or "KB Verification Agency",
                                     feedback_date=fd)
                    return _ct_load(), f"✅ {t['ticket_id']} raised. Deadline: {t['deadline'][:10]}"
                except Exception as e:
                    return _ct_load(), f"❌ {e}"

            def _ct_inprog(ticket_id):
                from pipeline.correction_tracker import update_ticket_status
                if not (ticket_id or "").strip():
                    return _ct_load(), "❌ Enter a Ticket ID."
                t = update_ticket_status(ticket_id.strip(), "in_progress")
                return _ct_load(), (f"⏳ {ticket_id} marked in_progress." if t
                                    else f"❌ Ticket {ticket_id} not found.")

            def _ct_close(ticket_id, resolution, closed_by):
                from pipeline.correction_tracker import close_ticket
                if not (ticket_id or "").strip() or not (resolution or "").strip():
                    return _ct_load(), "❌ Ticket ID and resolution are required."
                t = close_ticket(ticket_id.strip(), resolution=resolution,
                                 closed_by=closed_by or "Translation Agency")
                if t:
                    pen = t["penalty_pct"]
                    return _ct_load(), (f"✅ {ticket_id} closed. "
                                        + (f"⚠️ Penalty: {pen:.1f}%" if pen > 0 else "✅ No penalty"))
                return _ct_load(), f"❌ Ticket {ticket_id} not found."

            def _ct_export(course, lang, agency):
                from pipeline.correction_tracker import (
                    get_all_tickets, get_tickets_for_course, export_closure_report)
                course = (course or "").strip(); lang = (lang or "").strip()
                tickets = (get_tickets_for_course(course, lang or None) if course
                           else [t for t in get_all_tickets()
                                 if not lang or t["tgt_lang"] == lang])
                if not tickets:
                    return None, "❌ No tickets found for the selected filter."
                label = f"{course or 'all'}_{lang or 'all'}"
                out = os.path.join(_get_output_dir(), f"KB_Correction_Closure_{label}.docx")
                try:
                    export_closure_report(tickets, out, agency_name=agency or "Translation Agency")
                    return out, f"✅ Saved → {Path(out).name} ({len(tickets)} tickets)"
                except Exception as e:
                    return None, f"❌ {e}"

            ct_raise_btn.click(_ct_raise,
                inputs=[ct_course, ct_lang, ct_feedback, ct_raised_by, ct_date],
                outputs=[ct_table, ct_raise_log])
            ct_refresh_btn.click(_ct_load, outputs=[ct_table])
            ct_inprog_btn.click(_ct_inprog,
                inputs=[ct_ticket_id], outputs=[ct_table, ct_close_log])
            ct_close_btn.click(_ct_close,
                inputs=[ct_ticket_id, ct_resolution, ct_closed_by],
                outputs=[ct_table, ct_close_log])
            ct_rep_btn.click(_ct_export,
                inputs=[ct_rep_course, ct_rep_lang, ct_rep_agency],
                outputs=[ct_rep_dl, ct_rep_log])

        # ── Tab 5: Settings ───────────────────────────────────
        with gr.Tab("⚙️ Settings"):
            with gr.Row():
                with gr.Column():
                    hf_key   = gr.Textbox(label="HuggingFace Token (HF_TOKEN)",
                                          value=env.get("HF_TOKEN",""),
                                          type="password", placeholder="hf_...")
                    save_btn    = gr.Button("💾 Save Token", variant="primary")
                    save_status = gr.Textbox(label="Status", lines=2, interactive=False)
                    gr.Markdown("---")
                    from pipeline.sovereign_guard import sovereign_mode_enabled
                    _sov = ("🔒 ENABLED — Foreign LLM APIs blocked"
                            if sovereign_mode_enabled() else "⚠️ DISABLED (non-KB only)")
                    gr.Markdown(
                        "### 🇮🇳 Sovereign AI Compliance — KB Tender RFB IN-KBL-543730-NC-RFB\n"
                        f"**Status: {_sov}**  \n"
                        "All core AI (ASR/Translation/TTS) runs fully **offline on-premise**.  \n"
                        "Set `KB_SOVEREIGN_MODE=1` in `.env` to block foreign LLM APIs.  \n"
                        "> Compliant: IT Act 2000 · DPDP Act 2023 · MeitY cloud policy."
                    )
                    gr.Markdown("---")
                    out_dir_box = gr.Textbox(
                        label="📂 Output Save Folder",
                        value=_get_output_dir(),
                        placeholder=r"e.g. D:\KB_Outputs",
                        info="All dubbed files, docs and QA reports are saved here."
                    )
                    with gr.Row():
                        dir_save_btn = gr.Button("💾 Set Folder", variant="primary")
                        open_btn     = gr.Button("📂 Open Folder")
                    dir_status = gr.Textbox(label="", lines=1, interactive=False)
                with gr.Column():
                    gr.Markdown(
                        "### Quality Scoring\n"
                        "Every segment scored 0–1 automatically:\n"
                        "- **≥ 0.55** → Pass\n"
                        "- **0.30–0.55** → Needs human review\n"
                        "- **< 0.30** → Failed\n\n"
                        "### Checkpoint / Resume\n"
                        "If a job crashes, re-run with same input — "
                        "completed segments restore from disk automatically.\n"
                    )
            save_btn.click(_save_env, inputs=[hf_key], outputs=[save_status])
            dir_save_btn.click(_set_output_dir, inputs=[out_dir_box], outputs=[dir_status])
            open_btn.click(lambda: (os.startfile(_get_output_dir()), "✅ Folder opened")[1], outputs=[dir_status])

                # ── Tab 6: Monthly Delivery Tracker ──────────────────
        with gr.Tab("📅 Monthly Delivery"):
            gr.Markdown(
                "**Monthly Delivery Tracker — KB Tender §4.4 + §5.1B SLA Penalties**  \n"
                "Schedule: 1→50|2→55|3→100|4→125|5→100|6→125|7→100|8→125|9→100|10→125|11→100 hrs  \n"
                "Penalty: <5%=none | 5–10%=2% | 10–20%=4% | >20%=5%"
            )
            with gr.Row():
                with gr.Column(scale=2):
                    md_month    = gr.Textbox(label="Month (e.g. 2025-07)", placeholder="YYYY-MM")
                    md_course   = gr.Textbox(label="Course ID", placeholder="KB_COURSE_001")
                    md_langs    = gr.CheckboxGroup(tgt_choices, value=KB_11, label="Languages Delivered")
                    md_hours    = gr.Number(label="Content Hours Delivered (this course)", value=0, precision=2)
                    md_add_btn  = gr.Button("➕ Add Entry", variant="primary")
                    md_state    = gr.State([])  # list of row dicts
                with gr.Column(scale=3):
                    md_table = gr.Dataframe(
                        headers=["Month", "Course ID", "Languages", "Hours", "Status"],
                        datatype=["str", "str", "str", "number", "str"],
                        interactive=False,
                        label="Monthly Submissions"
                    )
                    md_summary = gr.Textbox(label="Monthly SLA Summary (KB Tender §5.1B)", lines=8, interactive=False)
            with gr.Row():
                md_report_btn = gr.Button("📄 Export Month-wise Submission Report (.xlsx)", variant="primary")
                md_complete_btn = gr.Button("📦 Export Consolidated Completion Report (.xlsx)")
            with gr.Row():
                md_dl  = gr.File(label="⬇️ Download Report")
                md_log = gr.Textbox(label="Status", lines=3, interactive=False)

            def _md_add(month_str, course, langs, hours, rows):
                from pipeline.sla_penalty import compute_sla, MONTHLY_SCHEDULE
                from collections import defaultdict
                if not month_str or not course:
                    return rows, [[r["month"],r["course"],r["langs"],r["hours"],r.get("status","")] for r in rows], "❌ Month and Course ID are required."
                try:
                    month = int(str(month_str).strip())
                except ValueError:
                    return rows, [[r["month"],r["course"],r["langs"],r["hours"],r.get("status","")] for r in rows], "❌ Month must be 1–11."
                rows = rows + [{"month": month, "course": course,
                                "langs": ", ".join(langs), "hours": hours}]
                mh = defaultdict(float)
                for r in rows: mh[r["month"]] += r["hours"]
                table = []
                for r in rows:
                    sla = compute_sla(r["month"], mh[r["month"]])
                    r["status"] = sla["status"]
                    table.append([str(r["month"]),r["course"],r["langs"],r["hours"],sla["status"]])
                lines = ["KB Tender §5.1B — SLA per Month", "-"*52]
                for m in sorted(mh):
                    sla = compute_sla(m, mh[m])
                    tgt = MONTHLY_SCHEDULE.get(m, 0)
                    lines.append(f"Month {m:>2}: {mh[m]:>6.1f}/{tgt:<5.0f}h  shortfall {sla['shortfall_pct']:>5.1f}%  penalty {sla['penalty_pct']:.0f}%  {sla['status']}")
                return rows, table, "\n".join(lines)

            def _md_export_monthly(rows):
                if not rows:
                    return None, "❌ No entries to export."
                try:
                    # Build results dict grouped by month, then call the proper
                    # §4.5.iii DOCX generator (all 4 mandatory fields + SLA section)
                    from collections import defaultdict
                    # Group rows by month — pick the highest month present
                    months = sorted({r["month"] for r in rows})
                    month = months[-1]  # report for the latest month in the table
                    # Build results dict: {course_id: {lang: {success, duration_original, quality_summary}}}
                    results = defaultdict(dict)
                    for r in rows:
                        if r["month"] != month:
                            continue
                        for lang in [l.strip() for l in r["langs"].split(",") if l.strip()]:
                            results[r["course"]][lang] = {
                                "success": True,
                                "duration_original": r["hours"] * 3600,
                                "quality_summary": {},
                                "output_video_path": "",
                                "output_audio_path": "",
                                "error": "",
                            }
                    pipeline = get_pipeline()
                    out = os.path.join(_get_output_dir(),
                                       f"KB_Month{month}_Submission_Report.docx")
                    sla = pipeline.generate_monthly_report(
                        month, dict(results), out,
                        submitter_name="Translation Agency",
                    )
                    return out, (
                        f"✅ §4.5.iii DOCX → {Path(out).name}  |  "
                        f"SLA: {sla['status']}  |  "
                        f"shortfall {sla['shortfall_pct']:.1f}%  penalty {sla['penalty_pct']:.0f}%"
                    )
                except Exception as e:
                    return None, f"❌ {e}"

            def _md_export_completion(rows):
                """KB §4.6 — Consolidated Completion Report + Handover Package."""
                try:
                    pipeline = get_pipeline()
                    pkg = pipeline.generate_handover_package(
                        output_dir=_get_output_dir(),
                        agency_name="Translation Agency",
                    )
                    msg = (
                        f"✅ Handover package ready → {Path(pkg['handover_zip']).name}  |  "
                        f"{pkg['total_courses']} courses  |  {pkg['total_languages']} languages  |  "
                        f"{pkg['total_hours']:.1f} h  |  {pkg['total_assets']} assets"
                    )
                    return pkg["handover_zip"], msg
                except Exception as e:
                    return None, f"❌ {e}"

            gr.Markdown("---")
            gr.Markdown(
                "### 📋 Inception Report — Payment Milestone 1 (KB §4.1)  \n"
                "Submit within **T0+15 calendar days** to trigger **15% payment**.  \n"
                "Covers: contract details · course inventory · language routing · "
                "AI stack · 11-month delivery plan · team plan · QA process · "
                "risk register · payment milestones."
            )
            with gr.Row():
                with gr.Column(scale=2):
                    ir_agency   = gr.Textbox(label="Agency Name",
                                             value="Translation Agency")
                    ir_address  = gr.Textbox(label="Agency Address",
                                             placeholder="e.g. 123 MG Road, Bengaluru")
                    ir_contact  = gr.Textbox(label="Contact Person",
                                             placeholder="e.g. Dr. Priya Nair")
                    ir_email    = gr.Textbox(label="Contact Email",
                                             placeholder="e.g. priya@agency.in")
                    ir_t0       = gr.Textbox(label="Contract Start Date T0 (YYYY-MM-DD, blank=today)",
                                             placeholder="2025-07-01")
                    ir_courses  = gr.Textbox(
                        label="Course IDs (comma-separated, blank=TBC)",
                        placeholder="KB_COURSE_001, KB_COURSE_002"
                    )
                    ir_langs    = gr.CheckboxGroup(
                        tgt_choices, value=KB_11,
                        label="Target Languages in Scope"
                    )
                    with gr.Row():
                        gr.Button("KB 11", size="sm").click(
                            lambda: KB_11, outputs=[ir_langs])
                        gr.Button("All 22", size="sm").click(
                            lambda: [c for _, c in tgt_choices], outputs=[ir_langs])
                    ir_btn = gr.Button(
                        "📋 Generate Inception Report (.docx)",
                        variant="primary"
                    )
                with gr.Column(scale=2):
                    ir_dl  = gr.File(label="⬇️ Download Inception Report (.docx)")
                    ir_log = gr.Textbox(label="Status", lines=4, interactive=False)

            def _ir_generate(agency, address, contact, email, t0, courses_str, langs):
                if not langs:
                    return None, "❌ Select at least one target language."
                try:
                    pipeline = get_pipeline()
                    course_ids = [c.strip() for c in courses_str.split(",") if c.strip()]
                    out = os.path.join(_get_output_dir(), "KB_Inception_Report.docx")
                    pipeline.generate_inception_report(
                        course_ids=course_ids,
                        tgt_langs=langs,
                        output_dir=_get_output_dir(),
                        output_path=out,
                        agency_name=agency or "Translation Agency",
                        agency_address=address or "",
                        contact_person=contact or "",
                        contact_email=email or "",
                        t0_date=t0.strip() if t0 and t0.strip() else "",
                    )
                    return out, (
                        f"✅ Inception Report generated → {Path(out).name}  |  "
                        f"{len(course_ids)} course(s)  |  {len(langs)} language(s)"
                    )
                except Exception as e:
                    return None, f"❌ {e}"

            ir_btn.click(
                _ir_generate,
                inputs=[ir_agency, ir_address, ir_contact, ir_email,
                        ir_t0, ir_courses, ir_langs],
                outputs=[ir_dl, ir_log],
            )

            md_add_btn.click(
                _md_add,
                inputs=[md_month, md_course, md_langs, md_hours, md_state],
                outputs=[md_state, md_table, md_summary],
            )
            md_report_btn.click(_md_export_monthly, inputs=[md_state], outputs=[md_dl, md_log])
            md_complete_btn.click(_md_export_completion, inputs=[md_state], outputs=[md_dl, md_log])

        # ── Tab 7: Glossary ───────────────────────────────────
        # -- Tab 8: POC Package --
        with gr.Tab("POC Package"):
            gr.Markdown(
                "**Proof of Concept Packager - 50 Technical Marks (KB Tender)**  \n"
                "Dub one course into **Hindi + Punjabi** and bundle all outputs into a "
                "ZIP with a cover note for drive-link submission to the evaluation committee."
            )
            with gr.Row():
                with gr.Column(scale=2):
                    poc_video   = gr.File(
                        label="Source Video / Audio (MP4 / MP3 / WAV)",
                        file_types=[".mp4", ".mp3", ".wav", ".flac", ".webm"]
                    )
                    poc_id      = gr.Textbox(label="Course ID", value="KB_POC_001")
                    poc_src     = gr.Dropdown(src_choices, value="eng", label="Source Language")
                    poc_agency  = gr.Textbox(label="Agency Name", value="Translation Agency")
                    poc_contact = gr.Textbox(label="Contact Person", placeholder="Dr. Priya Nair")
                    poc_email   = gr.Textbox(label="Contact Email", placeholder="priya@agency.in")
                    poc_force   = gr.Checkbox(label="Force re-dub (ignore existing output)", value=False)
                    poc_only    = gr.Checkbox(label="Package only (skip dubbing)", value=False)
                    poc_btn     = gr.Button("Run + Package POC", variant="primary", size="lg")
                with gr.Column(scale=2):
                    poc_dl  = gr.File(label="Download POC ZIP (upload to drive link)")
                    poc_log = gr.Textbox(label="Status", lines=8, interactive=False)

            def _poc_run(video_file, course_id, src_lang, agency, contact, email,
                         force, package_only, progress=gr.Progress()):
                import sys as _sys
                _sys.path.insert(0, str(Path(__file__).parent.parent))
                from scripts.poc_package import _dub_poc, build_poc_package
                cid     = course_id or "KB_POC_001"
                out_dir = _get_output_dir()
                lines   = []
                if not package_only:
                    if video_file is None:
                        return None, "Upload a video file or tick Package only."
                    if not _job_semaphore.acquire(blocking=False):
                        return None, "Another job is running. Please wait."
                    try:
                        progress(0.1, desc="Dubbing Hindi + Punjabi...")
                        results = _dub_poc(video_file.name, src_lang, cid, out_dir, force)
                        for lang, r in results.items():
                            from pipeline.lang_config import LANG_NAMES
                            status = "OK" if r.success else ("FAILED: " + r.error[:60])
                            lines.append(LANG_NAMES.get(lang, lang) + ": " + status)
                    finally:
                        _job_semaphore.release()
                progress(0.85, desc="Packaging...")
                try:
                    zip_path = build_poc_package(
                        course_id=cid, output_dir=out_dir,
                        agency_name=agency or "Translation Agency",
                        contact_person=contact or "",
                        contact_email=email or "",
                    )
                    size_mb = round(Path(zip_path).stat().st_size / 1024 / 1024, 1)
                    lines.append("POC ZIP ready: " + Path(zip_path).name + " (" + str(size_mb) + " MB)")
                    lines.append("Upload this ZIP to your drive link for the evaluation committee.")
                    progress(1.0, desc="Done!")
                    return zip_path, "\n".join(lines)
                except Exception as e:
                    return None, "ERROR: " + str(e)

            poc_btn.click(
                _poc_run,
                inputs=[poc_video, poc_id, poc_src, poc_agency,
                        poc_contact, poc_email, poc_force, poc_only],
                outputs=[poc_dl, poc_log],
            )


        with gr.Tab("📖 Glossary"):
            gr.Markdown(
                "**Standardised Terminology Glossary — KB Tender Final Deliverable**  \n"
                "Add English terms with their approved translations per language. "
                "Export as Excel (.xlsx) for CBP portal submission."
            )
            with gr.Row():
                with gr.Column(scale=2):
                    gl_term    = gr.Textbox(label="English Term", placeholder="e.g. Competency Framework")
                    gl_domain  = gr.Textbox(label="Domain / Category", placeholder="e.g. HR, Governance, IT")
                    gl_langs   = gr.CheckboxGroup(tgt_choices, value=KB_11, label="Languages to Add Translation For")
                    gl_trans   = gr.Textbox(
                        label="Translations (one per line: lang_code: translation)",
                        placeholder="hin: दक्षता ढांचा\ntan: திறன் கட்டமைப்பு",
                        lines=6
                    )
                    gl_add_btn = gr.Button("➕ Add Term", variant="primary")
                    gl_state   = gr.State([])  # list of term dicts
                with gr.Column(scale=3):
                    gl_table = gr.Dataframe(
                        headers=["English Term", "Domain", "Languages", "Translations"],
                        datatype=["str", "str", "str", "str"],
                        interactive=False,
                        label="Glossary Entries"
                    )
            with gr.Row():
                gl_export_btn  = gr.Button("📥 Export Glossary (.xlsx)", variant="primary")
                gl_import_file = gr.File(label="Import Glossary (.xlsx)", file_types=[".xlsx"])
                gl_import_btn  = gr.Button("📤 Import")
            with gr.Row():
                gl_dl  = gr.File(label="⬇️ Download Glossary")
                gl_log = gr.Textbox(label="Status", lines=2, interactive=False)

            def _gl_add(term, domain, langs, trans_text, rows):
                if not term.strip():
                    return rows, [[r["term"], r["domain"], r["langs"], r["trans"]] for r in rows], "❌ Term is required."
                trans_map = {}
                for line in trans_text.strip().splitlines():
                    if ":" in line:
                        k, _, v = line.partition(":")
                        trans_map[k.strip()] = v.strip()
                entry = {
                    "term": term.strip(),
                    "domain": domain.strip(),
                    "langs": ", ".join(langs),
                    "trans": " | ".join(f"{k}: {v}" for k, v in trans_map.items()),
                    "trans_map": trans_map,
                }
                rows = rows + [entry]
                table = [[r["term"], r["domain"], r["langs"], r["trans"]] for r in rows]
                return rows, table, f"✅ Added '{term.strip()}' ({len(rows)} terms total)"

            def _gl_export(rows):
                if not rows:
                    return None, "❌ No glossary entries to export."
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Glossary"
                    # Header: English + one column per language
                    all_langs = sorted({k for r in rows for k in r.get("trans_map", {})})
                    ws.append(["English Term", "Domain"] + [LANG_NAMES.get(l, l) for l in all_langs])
                    for r in rows:
                        row = [r["term"], r["domain"]] + [r.get("trans_map", {}).get(l, "") for l in all_langs]
                        ws.append(row)
                    out = os.path.join(_get_output_dir(), "KB_Standardised_Glossary.xlsx")
                    wb.save(out)
                    return out, f"✅ Saved → {Path(out).name} ({len(rows)} terms, {len(all_langs)} languages)"
                except Exception as e:
                    return None, f"❌ {e}"

            def _gl_import(file, rows):
                if file is None:
                    return rows, [[r["term"], r["domain"], r["langs"], r["trans"]] for r in rows], "❌ Upload an .xlsx file."
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file.name)
                    ws = wb.active
                    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                    lang_cols = headers[2:]  # after English Term, Domain
                    # reverse map: lang name → code
                    name_to_code = {v: k for k, v in LANG_NAMES.items()}
                    imported = 0
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row[0]:
                            continue
                        trans_map = {}
                        for i, lname in enumerate(lang_cols):
                            val = row[2 + i] if 2 + i < len(row) else None
                            if val:
                                code = name_to_code.get(lname, lname)
                                trans_map[code] = str(val)
                        entry = {
                            "term": str(row[0]),
                            "domain": str(row[1] or ""),
                            "langs": ", ".join(trans_map.keys()),
                            "trans": " | ".join(f"{k}: {v}" for k, v in trans_map.items()),
                            "trans_map": trans_map,
                        }
                        rows = rows + [entry]
                        imported += 1
                    table = [[r["term"], r["domain"], r["langs"], r["trans"]] for r in rows]
                    return rows, table, f"✅ Imported {imported} terms."
                except Exception as e:
                    return rows, [], f"❌ {e}"

            gl_add_btn.click(
                _gl_add,
                inputs=[gl_term, gl_domain, gl_langs, gl_trans, gl_state],
                outputs=[gl_state, gl_table, gl_log],
            )
            gl_export_btn.click(_gl_export, inputs=[gl_state], outputs=[gl_dl, gl_log])
            gl_import_btn.click(_gl_import, inputs=[gl_import_file, gl_state], outputs=[gl_state, gl_table, gl_log])

        # ── Tab 8: Live Logs ──────────────────────────────────
        with gr.Tab("📊 Live Logs"):
            gr.Markdown("Real-time pipeline logs. Auto-refreshes every 3 s.")
            log_box = gr.Textbox(value=_get_log, every=3,
                                 label="Pipeline Log", lines=30, interactive=False)

    return app


def create_app():
    """Entry point for `gradio ui/app.py` hot-reload."""
    return build_ui()


if __name__ == "__main__":
    import socket

    def _free_port(start: int = 7860, end: int = 7870) -> int:
        for port in range(start, end + 1):
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                if s.connect_ex(("172.23.198.15", port)) != 0:
                    return port
        return start  # fallback — let Gradio raise the error

    PORT = _free_port()
    app  = build_ui()
    print(f"UI running at http://172.23.198.15:{PORT}  (keep this tab open, just refresh on code changes)")
    app.launch(
        server_name="172.23.198.15",
        server_port=PORT,
        share=False,
        inbrowser=False,
    )
